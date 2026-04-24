import streamlit as st
import pandas as pd
import yfinance as yf
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Portfolio Tracker", layout="wide", page_icon="📈")

st.markdown("""
    <style>
    .main { padding: 1rem 2rem; }
    .stMetric { background: #f8f9fa; padding: 10px; border-radius: 8px; }
    .profit { color: #16a34a; font-weight: bold; }
    .loss { color: #dc2626; font-weight: bold; }
    </style>
""", unsafe_allow_html=True)

# ── FIFO Engine ─────────────────────────────────────────────────────────────

def normalize_ticker(ticker):
    ticker = str(ticker).strip().upper()
    return ticker[:-3] if ticker.endswith(".NS") else ticker


def to_nse_ticker(ticker):
    return f"{normalize_ticker(ticker)}.NS"


def compute_fifo_holdings(df_port):
    """Returns DataFrame of current holdings with FIFO avg cost."""
    queues = {}   # ticker -> list of [qty, price] lots
    for _, row in df_port.sort_values("Date").iterrows():
        ticker = row["Ticker"]
        qty    = float(row["Quantity"])
        price  = float(row["Price"])
        action = row["Action"].upper()

        if ticker not in queues:
            queues[ticker] = []

        if action == "BUY":
            queues[ticker].append([qty, price])

        elif action == "SELL":
            remaining = qty
            while remaining > 0 and queues[ticker]:
                lot_qty, lot_price = queues[ticker][0]
                if lot_qty <= remaining:
                    remaining -= lot_qty
                    queues[ticker].pop(0)
                else:
                    queues[ticker][0][0] -= remaining
                    remaining = 0

            if remaining > 0:
                available_qty = qty - remaining
                raise ValueError(
                    f"Invalid SELL for portfolio '{row['Portfolio']}', ticker '{ticker}' "
                    f"on {row['Date']:%d/%m/%y}: tried to sell {qty:g} shares, "
                    f"but only {available_qty:g} were available."
                )

    rows = []
    for ticker, lots in queues.items():
        total_qty   = sum(l[0] for l in lots)
        if total_qty <= 0:
            continue
        avg_cost    = sum(l[0] * l[1] for l in lots) / total_qty
        total_invested = sum(l[0] * l[1] for l in lots)
        rows.append({"Ticker": ticker, "Quantity": total_qty,
                     "Avg Cost (₹)": avg_cost, "Total Invested (₹)": total_invested})

    return pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["Ticker", "Quantity", "Avg Cost (₹)", "Total Invested (₹)"])


@st.cache_data(ttl=300, show_spinner=False)
def fetch_prices(tickers):
    prices = {}
    for t in tickers:
        try:
            info = yf.Ticker(t).fast_info
            prices[t] = round(info.last_price, 2)
        except Exception:
            prices[t] = None
    return prices


def enrich_holdings(holdings_df, prices):
    df = holdings_df.copy()
    df["Current Price (₹)"] = df["Ticker"].map(lambda ticker: prices.get(to_nse_ticker(ticker)))
    df["Market Value (₹)"]  = df["Quantity"] * df["Current Price (₹)"]
    df["Unrealized P&L (₹)"] = df["Market Value (₹)"] - df["Total Invested (₹)"]
    total_market_value = df["Market Value (₹)"].sum(skipna=True)
    if total_market_value > 0:
        df["Weight %"] = (df["Market Value (₹)"] / total_market_value) * 100
    else:
        df["Weight %"] = pd.NA
    return df

# ── Excel Export ─────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="1F3864")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
PROFIT_FONT   = Font(color="16A34A", bold=True)
LOSS_FONT     = Font(color="DC2626", bold=True)
ALT_FILL      = PatternFill("solid", start_color="EFF6FF")
BORDER_THIN   = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC")
)
COLS = ["Ticker", "Quantity", "Avg Cost (₹)", "Total Invested (₹)",
        "Current Price (₹)", "Market Value (₹)", "Weight %", "Unrealized P&L (₹)"]
COL_WIDTHS = [16, 12, 16, 20, 18, 18, 12, 20]

def style_cell(cell, value, col_name, row_idx):
    cell.border = BORDER_THIN
    cell.alignment = Alignment(horizontal="right" if col_name != "Ticker" else "left",
                               vertical="center")
    if row_idx % 2 == 0:
        cell.fill = ALT_FILL

    if col_name == "Unrealized P&L (₹)" and value is not None:
        try:
            cell.font = PROFIT_FONT if float(value) >= 0 else LOSS_FONT
        except (TypeError, ValueError):
            pass

    if col_name in ("Avg Cost (₹)", "Total Invested (₹)", "Current Price (₹)",
                    "Market Value (₹)", "Unrealized P&L (₹)"):
        cell.number_format = '#,##0.00'
    elif col_name == "Weight %":
        cell.number_format = '0.00"%"'
    elif col_name == "Quantity":
        cell.number_format = '#,##0'


def build_excel(all_holdings: dict) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for port_name, df in all_holdings.items():
        safe_name = port_name[:31]
        ws = wb.create_sheet(title=safe_name)

        # Title row
        ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
        title_cell = ws["A1"]
        title_cell.value = f"{port_name} — Holdings as of {datetime.today().strftime('%d %b %Y')}"
        title_cell.font  = Font(bold=True, size=12, color="1F3864")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # Header row
        for c_idx, col in enumerate(COLS, 1):
            cell = ws.cell(row=2, column=c_idx, value=col)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BORDER_THIN
        ws.row_dimensions[2].height = 20

        # Data rows
        for r_idx, row in enumerate(df[COLS].itertuples(index=False), start=3):
            for c_idx, (col, val) in enumerate(zip(COLS, row), 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                style_cell(cell, val, col, r_idx)
            ws.row_dimensions[r_idx].height = 18

        # Totals row
        total_row = r_idx + 1 if len(df) > 0 else 3
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = PatternFill("solid", start_color="DBEAFE")

        for c_idx, col in enumerate(COLS[1:], 2):
            cell = ws.cell(row=total_row, column=c_idx)
            cell.fill = PatternFill("solid", start_color="DBEAFE")
            if col in ("Total Invested (₹)", "Market Value (₹)", "Unrealized P&L (₹)"):
                col_letter = get_column_letter(c_idx)
                cell.value = f"=SUM({col_letter}3:{col_letter}{total_row-1})"
                cell.number_format = '#,##0.00'
                cell.font = Font(bold=True)
            elif col == "Weight %":
                cell.value = 100
                cell.number_format = '0.00"%"'
                cell.font = Font(bold=True)
        # Column widths
        for c_idx, width in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

        ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── UI ────────────────────────────────────────────────────────────────────────

st.title("📈 Equity Portfolio Tracker")
st.caption("Indian Stocks (NSE) · FIFO Cost Basis · Multi-Portfolio")

with st.sidebar:
    st.header("⚙️ Controls")
    uploaded = st.file_uploader("Upload Transactions CSV", type=["csv"],
                                help="CSV with columns: Portfolio, Date, Ticker, Action, Quantity, Price")
    st.markdown("---")
    st.markdown("**Expected CSV format:**")
    st.code("Portfolio,Date,Ticker,Action,Quantity,Price\nFamilyPortfolio,10/01/24,RELIANCE,BUY,50,2400", language="csv")
    st.markdown("- Enter the ticker without `.NS`\n- Action: `BUY` or `SELL`\n- Date: `DD/MM/YY`")
    st.markdown("---")
    sample_csv = """Portfolio,Date,Ticker,Action,Quantity,Price
FamilyPortfolio,10/01/23,RELIANCE,BUY,50,2400.00
FamilyPortfolio,15/03/23,INFY,BUY,100,1450.00
FamilyPortfolio,20/06/23,RELIANCE,BUY,25,2550.00
FamilyPortfolio,05/09/23,HDFCBANK,BUY,80,1600.00
FamilyPortfolio,12/01/24,INFY,SELL,40,1700.00
FamilyPortfolio,18/04/24,RELIANCE,SELL,20,2900.00
FundPortfolio,14/02/23,TCS,BUY,30,3200.00
FundPortfolio,22/05/23,WIPRO,BUY,200,420.00
FundPortfolio,10/08/23,ICICIBANK,BUY,60,950.00
FundPortfolio,30/11/23,TCS,BUY,20,3600.00
FundPortfolio,08/02/24,WIPRO,SELL,100,480.00
FundPortfolio,15/05/24,ICICIBANK,SELL,20,1100.00
RetirementPortfolio,01/06/22,HDFC,BUY,40,2600.00
RetirementPortfolio,15/09/22,KOTAKBANK,BUY,50,1850.00
RetirementPortfolio,20/01/23,SBIN,BUY,150,550.00
RetirementPortfolio,10/07/23,HDFC,SELL,15,2800.00
RetirementPortfolio,05/03/24,SBIN,BUY,100,720.00
"""
    st.download_button("📥 Download Sample CSV",
                       sample_csv,
                       file_name="sample_transactions.csv", mime="text/csv")

if uploaded is None:
    st.info("👈 Upload a transactions CSV from the sidebar to get started. You can download a sample file to see the expected format.")
    st.stop()

# Load & validate
try:
    df_raw = pd.read_csv(uploaded)
    df_raw.columns = [c.strip() for c in df_raw.columns]
    required_cols = {"Portfolio", "Date", "Ticker", "Action", "Quantity", "Price"}
    missing = required_cols - set(df_raw.columns)
    if missing:
        st.error(f"Missing columns: {missing}")
        st.stop()

    text_cols = ["Portfolio", "Ticker", "Action"]
    for col in text_cols:
        df_raw[col] = df_raw[col].map(lambda value: value.strip() if isinstance(value, str) else value)

    invalid_fields = {}
    for col in required_cols:
        missing_mask = df_raw[col].isna()
        if col in text_cols:
            missing_mask = missing_mask | (df_raw[col] == "")
        if missing_mask.any():
            invalid_fields[col] = (missing_mask[missing_mask].index + 2).tolist()

    if invalid_fields:
        details = "; ".join(
            f"{col} missing on rows {', '.join(map(str, rows[:5]))}{'...' if len(rows) > 5 else ''}"
            for col, rows in invalid_fields.items()
        )
        raise ValueError(f"Required fields cannot be blank. {details}")

    df_raw["Date"] = pd.to_datetime(df_raw["Date"], format="%d/%m/%y", errors="raise")
    df_raw["Ticker"] = df_raw["Ticker"].map(normalize_ticker)
except ValueError as e:
    st.error(str(e))
    st.stop()
except Exception as e:
    st.error(f"Error reading file: {e}. Date values must use DD/MM/YY format, for example 10/01/24.")
    st.stop()

portfolios  = sorted(df_raw["Portfolio"].unique())
all_tickers = df_raw["Ticker"].map(to_nse_ticker).unique().tolist()

# Fetch prices
with st.spinner("Fetching live NSE prices..."):
    prices = fetch_prices(all_tickers)

failed = [normalize_ticker(t) for t, p in prices.items() if p is None]
if failed:
    st.warning(f"Could not fetch prices for: {', '.join(failed)}. These will show as N/A.")

# Compute holdings for all portfolios
all_holdings = {}
for port in portfolios:
    df_port    = df_raw[df_raw["Portfolio"] == port]
    holdings   = compute_fifo_holdings(df_port)
    if not holdings.empty:
        holdings   = enrich_holdings(holdings, prices)
        all_holdings[port] = holdings

if not all_holdings:
    st.info("No open holdings found in the uploaded transactions.")
    st.caption("All positions may be fully sold, or the file may not contain any BUY transactions.")
    st.stop()

# ── Portfolio Tabs ────────────────────────────────────────────────────────────

tabs = st.tabs([f"📂 {p}" for p in all_holdings.keys()])

for tab, (port_name, df) in zip(tabs, all_holdings.items()):
    with tab:
        total_invested = df["Total Invested (₹)"].sum()
        df_priced      = df.dropna(subset=["Current Price (₹)"])
        total_mktval   = df_priced["Market Value (₹)"].sum()
        total_pnl      = df_priced["Unrealized P&L (₹)"].sum()
        total_pnl_pct  = (total_pnl / df_priced["Total Invested (₹)"].sum() * 100) if not df_priced.empty else 0

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Holdings", len(df))
        c2.metric("Total Invested", f"₹{total_invested:,.0f}")
        c3.metric("Market Value", f"₹{total_mktval:,.0f}" if total_mktval else "N/A")
        if total_mktval:
            pnl_str = f"₹{total_pnl:,.0f} ({total_pnl_pct:.1f}%)"
            c4.metric("Unrealized P&L", pnl_str, delta=f"{total_pnl_pct:.1f}%", delta_color="normal")
        else:
            c4.metric("Unrealized P&L", "N/A")

        st.markdown("---")

        # Keep numeric columns numeric so Streamlit sorting stays accurate.
        display_df = df[COLS].copy()

        def color_pnl(val):
            try:
                v = float(str(val).replace(",", "").replace("%", ""))
                return "color: #16a34a; font-weight:bold" if v >= 0 else "color: #dc2626; font-weight:bold"
            except Exception:
                return ""

        styled = display_df.style\
            .map(color_pnl, subset=["Unrealized P&L (₹)"])\
            .format({
                "Quantity": "{:,.0f}",
                "Avg Cost (₹)": "{:,.2f}",
                "Total Invested (₹)": "{:,.2f}",
                "Current Price (₹)": "{:,.2f}",
                "Market Value (₹)": "{:,.2f}",
                "Weight %": "{:.2f}%",
                "Unrealized P&L (₹)": "{:,.2f}",
            }, na_rep="N/A")\
            .set_properties(**{"text-align": "right"})\
            .set_properties(subset=["Ticker"], **{"text-align": "left"})

        st.dataframe(styled, use_container_width=True, hide_index=True)

        # Transaction history for this portfolio
        with st.expander("📋 View Transaction History"):
            txn_df = df_raw[df_raw["Portfolio"] == port_name].sort_values("Date", ascending=False)
            st.dataframe(txn_df.reset_index(drop=True), use_container_width=True, hide_index=True)

# ── Export ────────────────────────────────────────────────────────────────────

st.markdown("---")
st.subheader("📤 Export to Excel")
col1, col2 = st.columns([2, 5])
with col1:
    if st.button("Generate Excel Report", type="primary"):
        with st.spinner("Building Excel file..."):
            excel_bytes = build_excel(all_holdings)
        fname = f"portfolio_holdings_{datetime.today().strftime('%Y%m%d')}.xlsx"
        st.download_button("⬇️ Download Excel", data=excel_bytes,
                           file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
with col2:
    st.caption(f"Will export {len(all_holdings)} portfolio sheets · Live prices as of now · FIFO cost basis")
